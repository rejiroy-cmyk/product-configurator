#!/usr/bin/env python3
"""
matrix.py — Urinoir x Montagematerial cross-tab -> Urinoir-Matrix.csv + .xlsx

Same source as gen.js/pdf.js (cat-data.json: 51 Urinoirs, their element and the
CLEANED part list), rendered as a matrix instead of 51 sheets: Urinoirs on Y,
articles on X, X in the cell where they pair. The ENTFERNT positions live in
`removed` and are deliberately absent here -- `parts` is already the clean list.

CSV cannot hold an image, so the XLSX is the illustrated twin: 96 thumbnails,
from cat-data's own base64 map, else public/img, else ./img (the five articles
that are in no pool -- scraped with scrape-ch7-images.cjs, see README).

    python3 st-scraper/urinoir-sabag/matrix.py
"""
import json, re, base64, io, csv, os, glob

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, '..', '..'))
SRC  = os.path.join(HERE, 'cat-data.json')
EXTRA = os.path.join(HERE, 'img')
d = json.load(open(SRC))
rows, meta, img = d['rows'], d['meta'], d['img']

# ---- grouping: gen.js GROUP(), plus the 2141 298 correction below ----
MUST = {'3612 500', '3612 272'}
def group(b):
    if re.fullmatch(r'3612 40[234567]', b):        return 'element'
    if b == '3461 110' or b in MUST:               return 'immer'
    if b in ('3612 420', '3612 419'):              return 'wasser'
    if re.fullmatch(r'3451 1(72|73)', b):          return 'rohbau'
    if b.startswith('3451 '):                      return 'steuerung'
    if re.match(r'(3441|3431|3432|3421 24|8111|3471)', b): return 'ablauf'
    if b == '2141 298':                            return 'ablauf'  # Einlaufmanschette zu Urinoir Duravit
    if re.match(r'(8211|3962|3421 233)', b):       return 'befestigung'
    if re.match(r'(3482|2112 890)', b):            return 'trennwand'
    return 'weiteres'

GLABEL = {'element':'Installationselement', 'immer':'Montagematerial', 'wasser':'Wasseranschluss',
          'rohbau':'Rohbau-Set', 'steuerung':'Steuerung (Auswahl)', 'ablauf':'Ablauf & Einlauf',
          'befestigung':'Befestigung', 'trennwand':'Trennwand', 'weiteres':'Weiteres'}
ORDER = ['element','immer','wasser','rohbau','steuerung','ablauf','befestigung','trennwand','weiteres']

# ---- columns (x axis) ----
parts = set()
for r in rows: parts.update(r['parts'])
cols = []
for g in ORDER:
    cols += sorted(b for b in parts if group(b) == g)
assert len(cols) == len(parts)

# ---- rows (y axis) ----
rows = sorted(rows, key=lambda r: (r['series'].lower(), r['name'].lower(), r['sub'].lower()))
count = {b: sum(1 for r in rows if b in r['parts']) for b in cols}

def price(b):
    p = (meta.get(b) or {}).get('price')
    return '' if p is None else f'{p:.2f}'
def label(b):
    m = meta.get(b) or {}
    return m.get('label') or m.get('desc') or ''

# ---- thumbnails: cat-data base64 first, then local public/img ----
def thumb(b):
    v = img.get(b)
    if v and v.startswith('data:'):
        return base64.b64decode(v.split(',', 1)[1])
    base = '0' + b.replace(' ', '')
    hits = sorted(glob.glob(os.path.join(ROOT, 'public/img', f'*_{base}_*.webp')))
    if not hits:
        extra = os.path.join(EXTRA, base[1:] + '.webp')
        return open(extra, 'rb').read() if os.path.exists(extra) else None
    pref = [h for h in hits if '_100_000_' in h] or hits
    return open(pref[0], 'rb').read()

def main():
    # ---------------- CSV ----------------
    IDX = ['Serie','Art-Nr','Urinoir','Ausführung','GTIN','URP CHF','Installationselement']
    out = os.path.join(ROOT, 'Urinoir-Matrix.csv')
    with open(out, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f, delimiter=';')
        pad = ['']*(len(IDX)-1)
        w.writerow(['Gruppe']          + pad + [GLABEL[group(b)] for b in cols])
        w.writerow(['Art-Nr']          + pad + cols)
        w.writerow(['Bezeichnung']     + pad + [label(b) for b in cols])
        w.writerow(['URP CHF']         + pad + [price(b) for b in cols])
        w.writerow(['Anzahl Urinoirs'] + pad + [count[b] for b in cols])
        w.writerow(IDX + cols)
        for r in rows:
            w.writerow([r['series'], r['our'], r['name'], r['sub'], r['gtin'],
                        price(r['our']),
                        r['el'] or '— ohne Element (Wandmontage)']
                       + ['X' if b in r['parts'] else '' for b in cols])
    print('CSV ', out, f'{len(rows)} Urinoirs x {len(cols)} Artikel')

    # ---------------- XLSX ----------------
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage

    wb = Workbook(); ws = wb.active; ws.title = 'Matrix'
    GC = {'element':'FFE7D5','immer':'FFF3C4','wasser':'D8ECFB','rohbau':'E6E0F8','steuerung':'D6F0DC',
          'ablauf':'DDF2F4','befestigung':'F0E4D8','trennwand':'E9E9E9','weiteres':'F5F5F5'}
    thin = Side(style='thin', color='D0D0D0')
    box  = Border(left=thin, right=thin, top=thin, bottom=thin)
    NIDX = len(IDX) + 1                      # +1 for the picture column A
    FIRST = NIDX + 1                         # first part column
    HDR, IMG_R, ART_R, PRC_R, DATA = 1, 2, 3, 4, 5

    ws.cell(HDR, 1, 'Urinoir Montagekatalog · Matrix').font = Font(bold=True, size=12)
    ws.merge_cells(start_row=HDR, start_column=1, end_row=HDR, end_column=NIDX)

    c = FIRST
    for g in ORDER:
        n = sum(1 for b in cols if group(b) == g)
        if not n: continue
        cell = ws.cell(HDR, c, GLABEL[g])
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill('solid', fgColor=GC[g])
        if n > 1: ws.merge_cells(start_row=HDR, start_column=c, end_row=HDR, end_column=c+n-1)
        c += n

    for i, h in enumerate(IDX):
        cell = ws.cell(ART_R, i+2, h)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(vertical='bottom', wrap_text=True)
    ws.cell(ART_R, 1, 'Bild').font = Font(bold=True, size=9)

    tmp = []
    def put_img(blob, row, col, px=40):
        if not blob: return
        im = PILImage.open(io.BytesIO(blob)).convert('RGBA')
        bg = PILImage.new('RGBA', im.size, (255,255,255,255))
        bg.alpha_composite(im); im = bg.convert('RGB')
        im.thumbnail((px, px))
        buf = io.BytesIO(); im.save(buf, 'PNG'); buf.seek(0)
        x = XLImage(buf); x.width, x.height = im.size
        ws.add_image(x, f'{get_column_letter(col)}{row}')
        tmp.append(buf)

    for j, b in enumerate(cols):
        col = FIRST + j
        g = group(b)
        put_img(thumb(b), IMG_R, col)
        a = ws.cell(ART_R, col, b)
        a.font = Font(bold=True, size=8, name='Menlo')
        a.alignment = Alignment(textRotation=90, horizontal='center', vertical='bottom')
        p = ws.cell(PRC_R, col, price(b))
        p.font = Font(size=7, color='777777'); p.alignment = Alignment(horizontal='center')
        for r_ in (HDR, IMG_R, ART_R, PRC_R):
            ws.cell(r_, col).fill = PatternFill('solid', fgColor=GC[g])
    ws.cell(PRC_R, NIDX, 'URP CHF →').font = Font(size=7, italic=True, color='777777')
    ws.cell(PRC_R, NIDX).alignment = Alignment(horizontal='right')

    for i, r in enumerate(rows):
        R = DATA + i
        ws.row_dimensions[R].height = 34
        put_img(thumb(r['our']), R, 1, px=42)
        vals = [r['series'], r['our'], r['name'], r['sub'], r['gtin'], price(r['our']),
                r['el'] or '— ohne Element']
        for k, v in enumerate(vals):
            cell = ws.cell(R, k+2, v)
            cell.alignment = Alignment(vertical='center', wrap_text=(k in (2,3)))
            cell.font = Font(size=9, bold=(k == 1), color=('C0392B' if not r['el'] and k == 6 else '000000'))
        for j, b in enumerate(cols):
            cell = ws.cell(R, FIRST+j, 'X' if b in r['parts'] else '')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = box
            if b in r['parts']:
                cell.font = Font(bold=True, size=10, color='1B5E20')
                cell.fill = PatternFill('solid', fgColor=GC[group(b)])

    ws.column_dimensions['A'].width = 7
    for i, wdt in enumerate([13, 10, 26, 24, 15, 9, 11]):
        ws.column_dimensions[get_column_letter(i+2)].width = wdt
    for j in range(len(cols)):
        ws.column_dimensions[get_column_letter(FIRST+j)].width = 6.2
    ws.row_dimensions[IMG_R].height = 34
    ws.row_dimensions[ART_R].height = 62
    ws.freeze_panes = ws.cell(DATA, FIRST)

    # ---- Legende sheet ----
    lg = wb.create_sheet('Legende')
    lg.append(['Art-Nr','Gruppe','Bezeichnung','Beschreibung','URP CHF','Anzahl Urinoirs'])
    for cell in lg[1]: cell.font = Font(bold=True)
    for b in cols:
        m = meta.get(b) or {}
        lg.append([b, GLABEL[group(b)], label(b), m.get('desc',''), m.get('price'), count[b]])
    for col, wdt in zip('ABCDEF', [11, 20, 52, 70, 10, 14]):
        lg.column_dimensions[col].width = wdt
    lg.freeze_panes = 'A2'

    outx = os.path.join(ROOT, 'Urinoir-Matrix.xlsx')
    wb.save(outx)
    print('XLSX', outx)


if __name__ == '__main__':
    main()
